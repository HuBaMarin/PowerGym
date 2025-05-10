import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from datetime import datetime, timedelta

# Sprint data
data = {
    'Sprint': [
        'Primer Sprint', 'Primer Sprint', 'Primer Sprint',
        'Segundo Sprint', 'Segundo Sprint', 'Segundo Sprint', 'Segundo Sprint'
    ],
    'Tarea': [
        'Investigación Inicial', 'Planificación', 'Especificaciones',
        'Documentación Técnica', 'Manuales', 'Memoria del Proyecto', 'Cierre'
    ],
    'Fecha de inicio': [
        '24/03/2025', '31/03/2025', '07/04/2025', 
        '21/04/2025', '12/05/2025', '19/05/2025', '31/05/2025'
    ],
    'Duración (días)': [6, 6, 13, 20, 6, 6, 1],
    'Dias transcurridos': [0, 7, 14, 28, 49, 56, 68]
}

# Create DataFrame
df = pd.DataFrame(data)

# Calculate end dates
df['Fecha fin'] = df.apply(lambda row: 
    (datetime.strptime(row['Fecha de inicio'], '%d/%m/%Y') + 
     timedelta(days=row['Duración (días)'])).strftime('%d/%m/%Y'), 
    axis=1)

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Sprints"

# Add title
ws.merge_cells('A1:E1')
title_cell = ws['A1']
title_cell.value = "PLANIFICACIÓN DE SPRINTS"
title_cell.font = Font(bold=True, size=14)
title_cell.alignment = Alignment(horizontal='center')

# Add headers for first sprint
ws['A3'] = "PRIMER SPRINT"
ws.merge_cells('A3:E3')
ws['A3'].font = Font(bold=True)
ws['A3'].alignment = Alignment(horizontal='center')
ws['A3'].fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

# Headers
headers = ['Sprint', 'Tarea', 'Fecha de inicio', 'Duración (días)', 'Dias transcurridos']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

# Filter and add data for first sprint
primer_sprint = df[df['Sprint'] == 'Primer Sprint']
row_num = 5
for _, row in primer_sprint.iterrows():
    for col_num, value in enumerate(row, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal='center')
    row_num += 1

# Add headers for second sprint
ws.cell(row=row_num + 1, column=1).value = "SEGUNDO SPRINT"
ws.merge_cells(f'A{row_num + 1}:E{row_num + 1}')
ws.cell(row=row_num + 1, column=1).font = Font(bold=True)
ws.cell(row=row_num + 1, column=1).alignment = Alignment(horizontal='center')
ws.cell(row=row_num + 1, column=1).fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

# Headers for second sprint
row_num += 2
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=row_num, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

# Filter and add data for second sprint
segundo_sprint = df[df['Sprint'] == 'Segundo Sprint']
row_num += 1
for _, row in segundo_sprint.iterrows():
    for col_num, value in enumerate(row, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal='center')
    row_num += 1

# Add totals for each sprint
first_sprint_duration = primer_sprint['Duración (días)'].sum()
second_sprint_duration = segundo_sprint['Duración (días)'].sum()
total_duration = first_sprint_duration + second_sprint_duration

# First sprint total
primer_sprint_end = 5 + len(primer_sprint) - 1
ws.cell(row=primer_sprint_end + 1, column=3).value = "Total Primer Sprint:"
ws.cell(row=primer_sprint_end + 1, column=3).font = Font(bold=True)
ws.cell(row=primer_sprint_end + 1, column=4).value = first_sprint_duration
ws.cell(row=primer_sprint_end + 1, column=4).font = Font(bold=True)

# Second sprint total
segundo_sprint_end = row_num
ws.cell(row=segundo_sprint_end + 1, column=3).value = "Total Segundo Sprint:"
ws.cell(row=segundo_sprint_end + 1, column=3).font = Font(bold=True)
ws.cell(row=segundo_sprint_end + 1, column=4).value = second_sprint_duration
ws.cell(row=segundo_sprint_end + 1, column=4).font = Font(bold=True)

# Add grand total
ws.cell(row=segundo_sprint_end + 3, column=3).value = "DURACIÓN TOTAL DEL PROYECTO:"
ws.cell(row=segundo_sprint_end + 3, column=3).font = Font(bold=True)
ws.cell(row=segundo_sprint_end + 3, column=4).value = total_duration
ws.cell(row=segundo_sprint_end + 3, column=4).font = Font(bold=True)

# Create a new sheet for the chart
chart_sheet = wb.create_sheet(title="Gráfico")

# Create a more detailed table for chart data
chart_sheet['A1'] = "Sprint"
chart_sheet['B1'] = "Tarea"
chart_sheet['C1'] = "Fecha de inicio"
chart_sheet['D1'] = "Fecha fin"
chart_sheet['E1'] = "Duración (días)"
chart_sheet['F1'] = "Rango de fechas"

# Make headers bold
for col in range(1, 7):
    chart_sheet.cell(row=1, column=col).font = Font(bold=True)

# Copy data for the chart
row_idx = 2
for idx, row in df.iterrows():
    chart_sheet.cell(row=row_idx, column=1).value = row['Sprint']
    chart_sheet.cell(row=row_idx, column=2).value = row['Tarea']
    chart_sheet.cell(row=row_idx, column=3).value = row['Fecha de inicio']
    chart_sheet.cell(row=row_idx, column=4).value = row['Fecha fin']
    chart_sheet.cell(row=row_idx, column=5).value = row['Duración (días)']
    chart_sheet.cell(row=row_idx, column=6).value = f"{row['Fecha de inicio']} a {row['Fecha fin']}"
    row_idx += 1

# Create bar chart with horizontal bars
chart = BarChart()
chart.type = "bar"
chart.style = 10
chart.title = "Duración de Tareas por Sprint"
chart.x_axis.title = "Tarea"
chart.y_axis.title = "Duración (días)"

# Define data ranges
data = Reference(chart_sheet, min_col=5, min_row=3, max_row=row_idx-1)
cats = Reference(chart_sheet, min_col=2, min_row=4, max_row=row_idx-1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

# Add data labels with date ranges
chart.dataLabels = DataLabelList()
chart.dataLabels.showVal = True

# Use different colors for different sprints
for i, series in enumerate(chart.series):
    series.graphicalProperties.solidFill = "4472C4" if i == 0 else "ED7D31"

# Add the chart to the chart sheet
chart_sheet.add_chart(chart, "H2")

# Create separate columns for start and end points of task timeline
chart_sheet['J1'] = "Sprint/Tarea"
chart_sheet['K1'] = "Fecha de inicio (Días)"
chart_sheet['L1'] = "Duración (días)"
chart_sheet['M1'] = "Sprint"

# Convert dates to numbers of days from project start for visualization
project_start_date = datetime.strptime(df['Fecha de inicio'].min(), '%d/%m/%Y')  # First task start date

# Create a hierarchical organization of sprint names and task names
current_sprint = None
row_idx_timeline = 2

# Create a dedicated task list with sprints
chart_sheet['R1'] = "LISTA DE TAREAS POR SPRINT"
chart_sheet['R1'].font = Font(bold=True, size=12)
chart_sheet.merge_cells('R1:T1')
chart_sheet['R1'].alignment = Alignment(horizontal='center')
chart_sheet['R1'].fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

chart_sheet['R2'] = "Sprint"
chart_sheet['S2'] = "Tarea" 
chart_sheet['T2'] = "Duración (días)"
chart_sheet['R2'].font = Font(bold=True)
chart_sheet['S2'].font = Font(bold=True)
chart_sheet['T2'].font = Font(bold=True)
chart_sheet['R2'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
chart_sheet['S2'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
chart_sheet['T2'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# Add task and sprint data for clear reading
row_num = 3
for i in range(2, row_idx):
    sprint = chart_sheet.cell(row=i, column=1).value
    task = chart_sheet.cell(row=i, column=2).value
    duration = chart_sheet.cell(row=i, column=5).value
    
    chart_sheet.cell(row=row_num, column=18).value = sprint
    chart_sheet.cell(row=row_num, column=19).value = task
    chart_sheet.cell(row=row_num, column=20).value = duration
    
    row_num += 1

# Create a simplified bar chart with task names
task_chart = BarChart()
task_chart.type = "bar"
task_chart.style = 10
task_chart.title = "Duración de Tareas por Sprint"
task_chart.x_axis.title = "Tarea"
task_chart.y_axis.title = "Duración (días)"

# Define data for the task chart
task_data = Reference(chart_sheet, min_col=20, min_row=2, max_row=row_num-1)
task_cats = Reference(chart_sheet, min_col=19, min_row=3, max_row=row_num-1)
task_chart.add_data(task_data, titles_from_data=True)
task_chart.set_categories(task_cats)

# Color code by sprint
for i in range(0, row_num-3):
    sprint_type = chart_sheet.cell(row=i+3, column=18).value
    color = "4472C4" if "Primer" in sprint_type else "ED7D31"  # Blue for first sprint, orange for second
    task_chart.series[0].graphicalProperties.solidFill = color

# Add data labels
task_chart.dataLabels = DataLabelList()
task_chart.dataLabels.showVal = True

# Adjust chart size
task_chart.height = 15
task_chart.width = 20

# Add the task chart to the chart sheet
chart_sheet.add_chart(task_chart, "R10")

# Table for task names with sprints
chart_sheet['P1'] = "Task with Sprint"
chart_sheet['P1'].font = Font(bold=True)

for i in range(2, row_idx):
    sprint = chart_sheet.cell(row=i, column=1).value
    task = chart_sheet.cell(row=i, column=2).value
    
    # Add task with its sprint
    chart_sheet.cell(row=i, column=16).value = f"{task} ({sprint})"
    
    # If we have a new sprint, add it as a header row
    if sprint != current_sprint:
        # Add sprint header row
        chart_sheet.cell(row=row_idx_timeline, column=10).value = sprint
        chart_sheet.cell(row=row_idx_timeline, column=10).font = Font(bold=True)
        chart_sheet.cell(row=row_idx_timeline, column=11).value = None  # No value for sprint headers
        chart_sheet.cell(row=row_idx_timeline, column=12).value = None  # No value for sprint headers
        chart_sheet.cell(row=row_idx_timeline, column=13).value = sprint
        
        current_sprint = sprint
        row_idx_timeline += 1
    
    # Add the task with clear task name (no indentation)
    chart_sheet.cell(row=row_idx_timeline, column=10).value = task
    
    # Calculate days from project start to task start
    start_date = datetime.strptime(chart_sheet.cell(row=i, column=3).value, '%d/%m/%Y')
    days_from_start = (start_date - project_start_date).days
    chart_sheet.cell(row=row_idx_timeline, column=11).value = days_from_start
    
    # Duration remains the same
    chart_sheet.cell(row=row_idx_timeline, column=12).value = chart_sheet.cell(row=i, column=5).value
    
    # Sprint identifier
    chart_sheet.cell(row=row_idx_timeline, column=13).value = sprint
    
    # Add additional columns for enhanced chart labels
    chart_sheet.cell(row=row_idx_timeline, column=14).value = chart_sheet.cell(row=i, column=3).value  # Start date
    chart_sheet.cell(row=row_idx_timeline, column=15).value = chart_sheet.cell(row=i, column=4).value  # End date
    
    row_idx_timeline += 1

# Create stacked bar chart
stacked_chart = BarChart()
stacked_chart.type = "bar"
stacked_chart.style = 10
stacked_chart.title = "Cronograma de Tareas (Barras Apiladas)"
stacked_chart.x_axis.title = "Sprint/Tarea"
stacked_chart.y_axis.title = "Días desde inicio del proyecto"

# Define data ranges for stacked chart - skip the sprint header rows by filtering for non-zero values
dates_data = Reference(chart_sheet, min_col=11, min_row=3, max_row=row_idx_timeline-1)
duration_data = Reference(chart_sheet, min_col=12, min_row=3, max_row=row_idx_timeline-1)
cats = Reference(chart_sheet, min_col=10, min_row=3, max_row=row_idx_timeline-1)

# Add data series
stacked_chart.add_data(dates_data, titles_from_data=True)
stacked_chart.add_data(duration_data, titles_from_data=True)
stacked_chart.set_categories(cats)

# Configure as stacked bar chart
stacked_chart.grouping = "stacked"
stacked_chart.overlap = 100

# Format the series - first series (start dates) in light blue, second (duration) in darker blue
stacked_chart.series[0].graphicalProperties.solidFill = "B8CCE4"  # Light blue for start date
stacked_chart.series[1].graphicalProperties.solidFill = "4472C4"  # Darker blue for duration

# Add data labels
stacked_chart.dataLabels = DataLabelList()
stacked_chart.dataLabels.showVal = True

# Adjust the chart layout for better display of task names
stacked_chart.x_axis.tickLblPos = "low"  # Position labels at the bottom
stacked_chart.x_axis.tickLblSkip = 0     # Don't skip any labels
stacked_chart.height = 20                # Make the chart taller
stacked_chart.width = 30                 # Make the chart wider

# Add the stacked chart to the chart sheet
chart_sheet.add_chart(stacked_chart, "O2")

# Create second chart by sprint type
chart2 = BarChart()
chart2.type = "bar"
chart2.style = 12
chart2.title = "Duración Total por Sprint"
chart2.x_axis.title = "Sprint"
chart2.y_axis.title = "Duración (días)"

# Create summary table for second chart
chart_sheet['F1'] = "Sprint"
chart_sheet['G1'] = "Duración Total"
chart_sheet['F1'].font = Font(bold=True)
chart_sheet['G1'].font = Font(bold=True)

chart_sheet['F2'] = "Primer Sprint"
chart_sheet['G2'] = first_sprint_duration
chart_sheet['F3'] = "Segundo Sprint"
chart_sheet['G3'] = second_sprint_duration

# Define data for second chart
data2 = Reference(chart_sheet, min_col=7, min_row=1, max_row=3)
cats2 = Reference(chart_sheet, min_col=6, min_row=2, max_row=3)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)

# Add data labels to second chart
chart2.dataLabels = DataLabelList()
chart2.dataLabels.showVal = True

# Add the second chart to the chart sheet
chart_sheet.add_chart(chart2, "H30")

# Title for the chart sheet
chart_sheet.merge_cells('A1:I1')
chart_sheet['A1'].value = "GRÁFICOS DE PLANIFICACIÓN DE SPRINTS"
chart_sheet['A1'].font = Font(bold=True, size=14)
chart_sheet['A1'].alignment = Alignment(horizontal='center')

# Section titles for data tables
chart_sheet['A2'].value = "DATOS DE TAREAS"
chart_sheet['A2'].font = Font(bold=True)
chart_sheet.merge_cells('A2:F2')
chart_sheet['A2'].fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

chart_sheet['J2'].value = "DATOS PARA GRÁFICOS"
chart_sheet['J2'].font = Font(bold=True)
chart_sheet.merge_cells('J2:M2')
chart_sheet['J2'].fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

# Reposition headers for chart data tables
for col, header in zip(range(1, 7), ["Sprint", "Tarea", "Fecha de inicio", "Fecha fin", "Duración (días)", "Rango de fechas"]):
    chart_sheet.cell(row=3, column=col).value = header
    chart_sheet.cell(row=3, column=col).font = Font(bold=True)
    chart_sheet.cell(row=3, column=col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

for col, header in zip(range(10, 14), ["Sprint/Tarea", "Fecha de inicio (Días)", "Duración (días)", "Sprint"]):
    chart_sheet.cell(row=3, column=col).value = header
    chart_sheet.cell(row=3, column=col).font = Font(bold=True)
    chart_sheet.cell(row=3, column=col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# Shift the data rows down by 1 to account for the new header positions
for i in range(row_idx, 3, -1):
    for j in range(1, 14):
        if j <= 8 or j >= 10:  # Skip columns 9 (which is empty)
            src_cell = chart_sheet.cell(row=i-1, column=j)
            dst_cell = chart_sheet.cell(row=i, column=j)
            dst_cell.value = src_cell.value

row_idx += 1  # Adjust row index for the shift

# Update data references for charts
data = Reference(chart_sheet, min_col=5, min_row=3, max_row=row_idx-1)
cats = Reference(chart_sheet, min_col=2, min_row=4, max_row=row_idx-1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

# Update data references for stacked chart
dates_data = Reference(chart_sheet, min_col=11, min_row=3, max_row=row_idx_timeline-1)
duration_data = Reference(chart_sheet, min_col=12, min_row=3, max_row=row_idx_timeline-1)
cats = Reference(chart_sheet, min_col=10, min_row=3, max_row=row_idx_timeline-1)
stacked_chart.add_data(dates_data, titles_from_data=True)
stacked_chart.add_data(duration_data, titles_from_data=True)
stacked_chart.set_categories(cats)

# Auto-adjust column width for chart sheet
for column in chart_sheet.columns:
    max_length = 0
    column_letter = None
    for cell in column:
        try:
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    if column_letter:
        adjusted_width = (max_length + 2)
        chart_sheet.column_dimensions[column_letter].width = adjusted_width

# Auto-adjust column width
for column in ws.columns:
    max_length = 0
    column_letter = None
    for cell in column:
        try:
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    if column_letter:
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook
excel_file = "planificacion_sprints.xlsx"
wb.save(excel_file)

print(f"Excel file '{excel_file}' created successfully.")